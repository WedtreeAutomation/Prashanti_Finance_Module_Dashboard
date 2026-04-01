import pandas as pd
import re
import requests
​import streamlit as st
import os
import io
from datetime import datetime, timezone
from dotenv import load_dotenv
from azure.identity import ClientSecretCredential
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

# =============================
# ENVIRONMENT & CONFIG
# =============================
load_dotenv()

st.set_page_config(
    page_title="General Ledger Dashboard",
    page_icon="💠",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================
# CUSTOM CSS - CLEAN LIGHT THEME
# =============================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .stApp { background-color: #F0F4F8; }

    /* Hero Banner */
    .hero-banner {
        background: linear-gradient(135deg, #0F2044 0%, #1A3A6B 60%, #1D5FAA 100%);
        padding: 5rem 3rem; border-radius: 20px; color: white;
        box-shadow: 0 20px 60px rgba(15, 32, 68, 0.3);
        margin-bottom: 2rem; position: relative; overflow: hidden;
    }
    .hero-banner::before {
        content: ''; position: absolute; top: -50%; right: -10%;
        width: 500px; height: 500px;
        background: radial-gradient(circle, rgba(59,130,246,0.2) 0%, transparent 70%); border-radius: 50%;
    }
    .hero-banner h1 { color: white; font-size: 3.2rem; font-weight: 700; line-height: 1.15; margin-bottom: 1rem; letter-spacing: -0.5px; }
    .hero-banner p { font-size: 1.15rem; opacity: 0.8; color: #B8D4F5; max-width: 600px; }

    /* Feature Cards */
    .feature-card {
        background: white; padding: 2rem; border-radius: 16px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.06); height: 100%;
        border: 1px solid #E8EDF4; transition: transform 0.2s, box-shadow 0.2s;
    }
    .feature-card:hover { transform: translateY(-4px); box-shadow: 0 8px 30px rgba(0,0,0,0.1); }
    .feature-icon { font-size: 2rem; margin-bottom: 1rem; }
    .feature-title { font-weight: 700; color: #0F2044; margin-bottom: 0.5rem; font-size: 1.05rem; }
    .feature-text { color: #64748B; font-size: 0.9rem; line-height: 1.6; }

    /* Main Header */
    .main-header { font-size: 1.9rem; font-weight: 700; color: #0F2044; margin-bottom: 1.5rem; padding-bottom: 0.75rem; border-bottom: 2px solid #DCE5F0; letter-spacing: -0.3px; }

    /* KPI Cards */
    .kpi-container { padding: 1rem 1.2rem; border-radius: 16px; color: white; box-shadow: 0 8px 24px rgba(0,0,0,0.12); transition: transform 0.2s; border: 1px solid rgba(255,255,255,0.1); }
    .kpi-container:hover { transform: translateY(-4px); }
    .kpi-title { font-size: 0.7rem; font-weight: 600; opacity: 0.85; margin-bottom: 0.25rem; text-transform: uppercase; letter-spacing: 0.3px; white-space: nowrap; }
    .kpi-value { font-size: 1.2rem; font-weight: 700; line-height: 1.1; margin-bottom: 0.4rem; font-family: 'DM Mono', monospace; white-space: nowrap; }
    .kpi-sub { font-size: 0.68rem; font-weight: 500; background: rgba(255,255,255,0.18); padding: 2px 8px; border-radius: 20px; display: inline-block; }

    .bg-revenue { background: linear-gradient(135deg, #0AB370 0%, #059669 100%); }
    .bg-expense { background: linear-gradient(135deg, #F43F5E 0%, #BE123C 100%); }
    .bg-profit { background: linear-gradient(135deg, #2563EB 0%, #1D4ED8 100%); }
    .bg-margin { background: linear-gradient(135deg, #7C3AED 0%, #5B21B6 100%); }

    /* Sidebar Theme */
    section[data-testid="stSidebar"] { background-color: #FFFFFF !important; border-right: 1px solid #E8EDF4; }
    section[data-testid="stSidebar"] .stApp { background-color: transparent; }
    section[data-testid="stSidebar"] p, section[data-testid="stSidebar"] span { color: #475569 !important; }
    
    section[data-testid="stSidebar"] div[data-baseweb="select"] *,
    section[data-testid="stSidebar"] input,
    section[data-testid="stSidebar"] div[class*="stSelectbox"] *,
    section[data-testid="stSidebar"] div[class*="stNumberInput"] * { color: initial !important; }
    
    section[data-testid="stSidebar"] label p { color: #334155 !important; font-weight: 600; font-size: 0.85rem; }
    section[data-testid="stSidebar"] h2 { color: #0F2044 !important; font-weight: 700; font-size: 1.3rem; letter-spacing: -0.3px; border-bottom: 2px solid #E8EDF4; padding-bottom: 10px; margin-bottom: 10px; }
    section[data-testid="stSidebar"] hr { border-color: #E8EDF4; margin: 15px 0; }

    section[data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"] {
        background: linear-gradient(135deg, #0F2044 0%, #1D5FAA 100%); color: white !important; border: none; border-radius: 8px; font-weight: 600; transition: all 0.2s;
    }
    section[data-testid="stSidebar"] .stButton button[data-testid="baseButton-primary"]:hover { transform: translateY(-2px); box-shadow: 0 4px 12px rgba(15, 32, 68, 0.2); }
    section[data-testid="stSidebar"] .stButton button[data-testid="baseButton-secondary"] { background: #F8FAFC; color: #0F2044 !important; border: 1px solid #CBD5E1; border-radius: 8px; font-weight: 600; }

    .sidebar-user-info {
        background: #F8FAFC; padding: 12px 15px; border-radius: 12px; margin-bottom: 15px;
        border-left: 4px solid #1D5FAA; border-top: 1px solid #E8EDF4; border-right: 1px solid #E8EDF4; border-bottom: 1px solid #E8EDF4;
    }
</style>
""", unsafe_allow_html=True)

# =============================
# AUTHENTICATION & API LOGIC
# =============================
@st.cache_resource
def get_credential():
    return ClientSecretCredential(
        tenant_id=os.getenv("FABRIC_TENANT_ID"),
        client_id=os.getenv("FABRIC_CLIENT_ID"),
        client_secret=os.getenv("FABRIC_CLIENT_SECRET")
    )

def get_access_token():
    credential = get_credential()
    token = credential.get_token("https://api.fabric.microsoft.com/.default")
    return token.token

def run_graphql(query, variables=None):
    endpoint = os.getenv("FABRIC_ENDPOINT")
    headers = {
        "Authorization": f"Bearer {get_access_token()}",
        "Content-Type": "application/json"
    }
    payload = {"query": query, "variables": variables}
    try:
        response = requests.post(endpoint, json=payload, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        st.error(f"API Error: {str(e)}")
        if hasattr(e, 'response') and e.response is not None:
            st.json(e.response.text)
        raise e

# =============================
# BRAND BASED QUERIES
# =============================
if st.session_state.get("brand") == "wed":

    READ_QUERY = """
    query {
      executesp_wd_readData { id Ledger classification ContraName Store Balance Year MonthName Month FinancialYearMonth last_modified_at last_modified_user }
    }
    """

    UPSERT_MUTATION = """
    mutation upsertBalance($year: Int!, $monthName: String!, $store: String!, $balance: Float, $account_name: String!, $classification: String!, $partner_id_name: String!, $last_modified_at: DateTime, $last_modified_user: String) {
      executesp_wd_upsertBalance(year: $year, monthName: $monthName, store: $store, balance: $balance, account_name: $account_name, classification: $classification, partner_id_name: $partner_id_name, last_modified_at: $last_modified_at, last_modified_user: $last_modified_user) { rows_affected }
    }
    """

else:  # PRA (default)

    READ_QUERY = """
    query {
      executesp_pr_readData { id account_name classification partner_id_name Store Balance Year MonthName Month FinancialYearMonth last_modified_at last_modified_user }
    }
    """

    UPSERT_MUTATION = """
    mutation upsertBalance($year: Int!, $monthName: String!, $store: String!, $balance: Float, $account_name: String!, $classification: String!, $partner_id_name: String!, $last_modified_at: DateTime, $last_modified_user: String) {
      executesp_pr_upsertBalance(year: $year, monthName: $monthName, store: $store, balance: $balance, account_name: $account_name, classification: $classification, partner_id_name: $partner_id_name, last_modified_at: $last_modified_at, last_modified_user: $last_modified_user) { rows_affected }
    }
    """

# =============================
# HELPER FUNCTIONS
# =============================
def get_financial_year_range(df, year, start_month=4):
    months = [(year, m) for m in range(start_month, 13)] + [(year + 1, m) for m in range(1, start_month)]
    periods = []
    for y, m in months:
        period_sort = f"{y}-{str(m).zfill(2)}"
        mask = (df['Year'] == y) & (df['Month'] == m)
        if mask.any():
            display = df.loc[mask, 'DisplayPeriod'].iloc[0]
            periods.append({'sort': period_sort, 'display': display})
    return periods

def calculate_profit_metrics(df, periods, revenue_classes):
    results = []
    for period in periods:
        period_data = df[df['DisplayPeriod'] == period]
        revenue = period_data[period_data['classification'].isin(revenue_classes)]['Balance'].sum()
        expenses = period_data[~period_data['classification'].isin(revenue_classes)]['Balance'].sum()
        profit = revenue - expenses
        results.append({
            'DisplayPeriod': period,
            'Revenue': revenue,
            'Expenses': expenses,
            'Profit': profit,
            'Margin': (profit / revenue * 100) if revenue != 0 else 0
        })
    return pd.DataFrame(results)

def fmt_currency(val):
    if pd.isna(val) or val == 0: 
        return "₹0"
    
    # 1. Format as integer string (No decimals)
    main_part = f"{int(round(float(val)))}"
    
    # 2. Logic for Indian Thousand Separator (Lakhs/Crores)
    if len(main_part) > 3:
        # Separate the last 3 digits
        last_three = main_part[-3:]
        remaining = main_part[:-3]
        
        # Group the remaining digits in pairs (twos)
        remaining = re.sub(r'(\d+?)(?=(\d{2})+(?!\d))', r'\1,', remaining)
        main_part = remaining + ',' + last_three
        
    return f"₹{main_part}"

# =============================
# HIERARCHICAL DATA BUILDER
# =============================
def build_hierarchy_data(report_df, grouping_list, group_by='DisplayPeriod'):
    """
    Modified to support dynamic grouping (e.g., by Period or by Store).
    
    :param report_df: The source DataFrame
    :param grouping_list: List of columns to show (e.g., selected_periods or unique_stores)
    :param group_by_col: The column name in report_df to filter against (default 'DisplayPeriod')
    """
    hierarchy = {}
    
    classification_order = [
        'Net Sales', 'Other Income', 'Cost of Goods Sold (COGS)', 'Employee cost',
        'Rent and Utilities', 'Marketing and Advertisment', 'Admin Expenses',
        'Logistics', 'Other Expenses', 'Finance cost', 'Supplier Payments',
        'Purchase Expense', 'Depreciation'
    ]
    
    unique_classifications = report_df['classification'].dropna().unique()
    
    def get_classification_order(cls):
        try:
            return classification_order.index(cls)
        except ValueError:
            return len(classification_order) + (sum(ord(c) for c in cls) if cls else 0)
    
    sorted_classifications = sorted(unique_classifications, key=get_classification_order)
    
    for classification in sorted_classifications:
        cls_df = report_df[report_df['classification'] == classification]
        cls_totals = {}
        for item in grouping_list:
            # Dynamically filter by the provided column name
            cls_totals[item] = cls_df[cls_df[group_by] == item]['Balance'].sum()
        
        accounts = {}
        for account in sorted(cls_df['account_name'].dropna().unique()):
            acc_df = cls_df[cls_df['account_name'] == account]
            acc_totals = {}
            for item in grouping_list:
                acc_totals[item] = acc_df[acc_df[group_by] == item]['Balance'].sum()
            
            partners = {}
            for partner in sorted(acc_df['partner_id_name'].dropna().unique()):
                prt_df = acc_df[acc_df['partner_id_name'] == partner]
                prt_totals = {}
                for item in grouping_list:
                    prt_totals[item] = prt_df[prt_df[group_by] == item]['Balance'].sum()
                partners[partner] = prt_totals
            
            accounts[account] = {'totals': acc_totals, 'partners': partners}
        
        hierarchy[classification] = {'totals': cls_totals, 'accounts': accounts}
        
    return hierarchy

# =============================
# EXCEL EXPORT WITH NATIVE GROUPING
# =============================
def build_excel_report(hierarchy, periods, store_filter="All", brand="pra", report_type="pnl", expand_all=False, open_classifications=None, open_accounts=None):
    brand_name = "Prashanti" if brand == "pra" else "Wedtree"
    if open_classifications is None: open_classifications = set()
    if open_accounts is None: open_accounts = set()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Statement"

    ws.sheet_properties.outlinePr.summaryBelow = False
    
    DARK_BLUE, MID_BLUE, LIGHT_BLUE, WHITE = "0F2044", "1A3A6B", "EEF2F9", "FFFFFF"
    GREEN, RED, GREY = "059669", "E11D48", "64748B"
    
    def make_fill(hex_color): return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    hair_border = Border(bottom=Side(style='hair', color='E2E8F0'))
    
    # --- Title Block ---
    ws.merge_cells(f"A1:{get_column_letter(len(periods) + 3)}1")
    title_cell = ws["A1"]
    if report_type == "store":
        title_cell.value = "Store Comparison Report"
    else:
        title_cell.value = "Profit & Loss Statement"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title_cell.fill = make_fill(DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells(f"A2:{get_column_letter(len(periods) + 3)}2")
    sub_cell = ws["A2"]
    if report_type == "store":
        sub_cell.value = f"Brand: {brand_name}  |  Comparison View  |  Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    else:
        sub_cell.value = f"Brand: {brand_name}  |  Store: {store_filter}  |  Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    sub_cell.font = Font(name="Calibri", size=9, color="B8D4F5")
    sub_cell.fill = make_fill(MID_BLUE)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6
    
    # --- Headers ---
    header_row = 4
    headers = ["Account Hierarchy"] + list(periods) + ["Total"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left", vertical="center", wrap_text=False)
        cell.border = Border(bottom=Side(style='medium', color='FFFFFF'))
    ws.row_dimensions[header_row].height = 24
    
    ws.column_dimensions['A'].width = 38
    for col_idx in range(2, len(periods) + 2): ws.column_dimensions[get_column_letter(col_idx)].width = 15
    ws.column_dimensions[get_column_letter(len(periods) + 2)].width = 15
    
    # --- Data Population with Outline Grouping ---
    current_row = header_row + 1
    
    classification_order = [
        'Net Sales',
        'Other Income',
        'Cost of Goods Sold (COGS)',
        'Employee cost',
        'Rent and Utilities',            
        'Marketing and Advertisment',    
        'Admin Expenses',
        'Logistics',
        'Other Expenses',
        'Finance cost',
        'Supplier Payments',            
        'Purchase Expense',
        'Depreciation'
    ]
    
    def get_classification_order(cls):
        try:
            return classification_order.index(cls)
        except ValueError:
            return len(classification_order) + (sum(ord(c) for c in cls) if cls else 0)
    
    sorted_classifications = sorted(hierarchy.keys(), key=get_classification_order)
    
    for cls_name in sorted_classifications:
        cls_data = hierarchy[cls_name]
        cls_key = f"cls_{cls_name}"
        is_cls_open = expand_all or (cls_key in open_classifications)
        cls_total = sum(cls_data['totals'].values())
        
        # 1. Classification (Level 0 summary - no outline level)
        cls_cell = ws.cell(row=current_row, column=1, value=f"  {cls_name}")
        cls_cell.font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
        cls_cell.fill = make_fill(LIGHT_BLUE)
        cls_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cls_cell.border = Border(top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))
        
        for col_idx, p in enumerate(periods, 2):
            v = cls_data['totals'].get(p, 0)
            cell = ws.cell(row=current_row, column=col_idx, value=v)
            cell.number_format = '#,##0'
            cell.font = Font(name="Calibri", bold=True, size=9, color=GREEN if v > 0 else RED if v < 0 else GREY)
            cell.fill = make_fill(LIGHT_BLUE)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))
        
        total_cell = ws.cell(row=current_row, column=len(periods) + 2, value=cls_total)
        total_cell.number_format = '#,##0'
        total_cell.font = Font(name="Calibri", bold=True, size=9, color=GREEN if cls_total > 0 else RED if cls_total < 0 else GREY)
        total_cell.fill = make_fill(LIGHT_BLUE)
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.border = Border(top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        # Sort accounts alphabetically within each classification
        for acc_name in sorted(cls_data['accounts'].keys()):
            acc_data = cls_data['accounts'][acc_name]
            acc_key = f"acc_{cls_name}__{acc_name}"
            is_acc_open = expand_all or (acc_key in open_accounts)
            acc_total = sum(acc_data['totals'].values())
            
            # 2. Account (Level 1 details inside Classification)
            acc_cell = ws.cell(row=current_row, column=1, value=f"      {acc_name}")
            acc_cell.font = Font(name="Calibri", bold=True, size=9, color="334155")
            acc_cell.fill = make_fill(WHITE)
            acc_cell.alignment = Alignment(horizontal="left", vertical="center", indent=3)
            acc_cell.border = hair_border
            
            for col_idx, p in enumerate(periods, 2):
                v = acc_data['totals'].get(p, 0)
                cell = ws.cell(row=current_row, column=col_idx, value=v)
                cell.number_format = '#,##0'
                cell.font = Font(name="Calibri", bold=True, size=8, color=GREEN if v > 0 else RED if v < 0 else GREY)
                cell.fill = make_fill(WHITE)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = hair_border
            
            acc_total_cell = ws.cell(row=current_row, column=len(periods) + 2, value=acc_total)
            acc_total_cell.number_format = '#,##0'
            acc_total_cell.font = Font(name="Calibri", bold=True, size=8, color=GREEN if acc_total > 0 else RED if acc_total < 0 else GREY)
            acc_total_cell.fill = make_fill(WHITE)
            acc_total_cell.alignment = Alignment(horizontal="right", vertical="center")
            acc_total_cell.border = hair_border
            
            # ** EXCEL NATIVE GROUPING LOGIC **
            ws.row_dimensions[current_row].outline_level = 1
            if not is_cls_open:
                ws.row_dimensions[current_row].hidden = True
                
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
            # Sort partners alphabetically within each account
            for partner_name in sorted(acc_data['partners'].keys()):
                prt_totals = acc_data['partners'][partner_name]
                prt_total = sum(prt_totals.values())
                
                # 3. Partner (Level 2 details inside Account)
                prt_cell = ws.cell(row=current_row, column=1, value=f"            · {partner_name}")
                prt_cell.font = Font(name="Calibri", size=8, color=GREY)
                prt_cell.fill = make_fill(WHITE)
                prt_cell.alignment = Alignment(horizontal="left", vertical="center", indent=5)
                prt_cell.border = hair_border
                
                for col_idx, p in enumerate(periods, 2):
                    v = prt_totals.get(p, 0)
                    cell = ws.cell(row=current_row, column=col_idx, value=v)
                    cell.number_format = '#,##0'
                    cell.font = Font(name="Calibri", size=8, color=GREEN if v > 0 else RED if v < 0 else GREY)
                    cell.fill = make_fill(WHITE)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = hair_border
                
                prt_total_cell = ws.cell(row=current_row, column=len(periods) + 2, value=prt_total)
                prt_total_cell.number_format = '#,##0'
                prt_total_cell.font = Font(name="Calibri", size=8, color=GREEN if prt_total > 0 else RED if prt_total < 0 else GREY)
                prt_total_cell.fill = make_fill(WHITE)
                prt_total_cell.alignment = Alignment(horizontal="right", vertical="center")
                prt_total_cell.border = hair_border
                
                # ** EXCEL NATIVE GROUPING LOGIC **
                ws.row_dimensions[current_row].outline_level = 2
                if not is_cls_open or not is_acc_open:
                    ws.row_dimensions[current_row].hidden = True
                    
                ws.row_dimensions[current_row].height = 16
                current_row += 1
    
    # Grand Total
    grand_totals = {p: sum(cls_data['totals'].get(p, 0) for cls_data in hierarchy.values()) for p in periods}
    grand_total = sum(grand_totals.values())
    
    gt_cell = ws.cell(row=current_row, column=1, value="  GRAND TOTAL")
    gt_cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    gt_cell.fill = make_fill(DARK_BLUE)
    gt_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    for col_idx, p in enumerate(periods, 2):
        cell = ws.cell(row=current_row, column=col_idx, value=grand_totals[p])
        cell.number_format = '#,##0'
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = make_fill(DARK_BLUE)
        cell.alignment = Alignment(horizontal="right", vertical="center")
    
    gt_total = ws.cell(row=current_row, column=len(periods) + 2, value=grand_total)
    gt_total.number_format = '#,##0'
    gt_total.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    gt_total.fill = make_fill(DARK_BLUE)
    gt_total.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[current_row].height = 24
    
    ws.freeze_panes = "B5"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# =============================
# PERSISTENT SESSION STATE
# =============================
if "logged_in" not in st.session_state:
    if st.query_params.get("auth") == "active":
        st.session_state.logged_in = True
        st.session_state.logged_in_user = st.query_params.get("user", "")
    else:
        st.session_state.logged_in = False
        st.session_state.logged_in_user = ""

if "open_classifications" not in st.session_state: st.session_state.open_classifications = set()
if "open_accounts" not in st.session_state: st.session_state.open_accounts = set()
if "reset_editor" not in st.session_state: st.session_state.reset_editor = 0

# New session state for global HTML table expand/collapse
if "expand_all_mode" not in st.session_state: 
    st.session_state.expand_all_mode = False

params = st.query_params
if "toggle_cls" in params:
    key = params["toggle_cls"]
    if key in st.session_state.open_classifications:
        st.session_state.open_classifications.discard(key)
        st.session_state.open_accounts = {a for a in st.session_state.open_accounts if not a.startswith(f"acc_{key.replace('cls_', '')}__")}
    else:
        st.session_state.open_classifications.add(key)
    del st.query_params["toggle_cls"]
    st.rerun()

if "toggle_acc" in params:
    key = params["toggle_acc"]
    if key in st.session_state.open_accounts: st.session_state.open_accounts.discard(key)
    else: st.session_state.open_accounts.add(key)
    del st.query_params["toggle_acc"]
    st.rerun()

# New separate session states for independent table control
if "expand_pnl" not in st.session_state: 
    st.session_state.expand_pnl = False

if "expand_store" not in st.session_state: 
    st.session_state.expand_store = False

# =============================
# APP FLOW: LOGGED OUT
# =============================
if not st.session_state.logged_in:
    with st.sidebar:
        st.markdown("<h2 style='color:#0F2044; font-weight:700;'>Secure Login</h2>", unsafe_allow_html=True)
        st.write("Sign in to access the dashboard.")
        with st.form("login_form"):
            username = st.text_input("Email", placeholder="name@company.com")
            password = st.text_input("Password", type="password", placeholder="••••••••")
            submitted = st.form_submit_button("Sign In →", width = 'stretch', type="primary")
            if submitted:
                if username == os.getenv("APP_USERNAME") and password == os.getenv("APP_PASSWORD"):
                    st.session_state.logged_in = True
                    st.session_state.logged_in_user = username
                    st.query_params["auth"] = "active"
                    st.query_params["user"] = username
                    st.rerun()
                else:
                    st.error("Invalid credentials.")

    st.markdown("""
    <div class="hero-banner">
        <h1>Financial Intelligence Hub.</h1>
        <p>Your centralized dashboard for general ledger management, powerful financial reporting, and real-time data adjustments.</p>
    </div>
    """, unsafe_allow_html=True)

    cols = st.columns(3)
    features = [
        ("✨", "Real-time Integration", "Synchronize instantly with Microsoft Fabric for single-source-of-truth accuracy."),
        ("📊", "Dynamic Insights", "Visualize P&L comparisons, custom periods, and hierarchical data instantly."),
        ("✏️", "Inline Editing", "Securely update balance adjustments with full audit tracking directly in the app.")
    ]
    for col, (icon, title, text) in zip(cols, features):
        col.markdown(f"""
        <div class="feature-card">
            <div class="feature-icon">{icon}</div>
            <div class="feature-title">{title}</div>
            <div class="feature-text">{text}</div>
        </div>
        """, unsafe_allow_html=True)
    st.stop()

# =============================
# BRAND SELECTION (AFTER LOGIN)
# =============================
if "brand" not in st.session_state:
    st.session_state.brand = "pra"

brand_map = {
    "Prashanti": "pra",
    "Wedtree": "wed"
}

reverse_brand_map = {v: k for k, v in brand_map.items()}

# with st.sidebar:
#     st.markdown("<h2 style='color:#0F2044; font-weight:700;'>Brand</h2>", unsafe_allow_html=True)

#     selected_label = st.selectbox(
#         "Select Brand",
#         options=list(brand_map.keys()),
#         index=list(brand_map.keys()).index(reverse_brand_map[st.session_state.brand])
#     )

#     selected_value = brand_map[selected_label]

#     if selected_value != st.session_state.brand:
#         st.session_state.brand = selected_value
#         st.rerun()

# =============================
# DATA FETCHING
# =============================
@st.cache_data(ttl=300)
def load_data(brand):
    result = run_graphql(READ_QUERY)
    data_key = "executesp_wd_readData" if brand == "wed" else "executesp_pr_readData"
    items = result.get("data", {}).get(data_key, [])
    df = pd.DataFrame(items)
    if brand == "wed":
        df = df.rename(columns={
            "Ledger": "account_name",
            "ContraName": "partner_id_name"
        })
    if not df.empty:
        df['Balance'] = pd.to_numeric(df['Balance'], errors='coerce').fillna(0.0)
        df['Year'] = pd.to_numeric(df['Year'], errors='coerce')
        df['Month'] = pd.to_numeric(df['Month'], errors='coerce')
        df['PeriodSort'] = df['Year'].astype(str) + "-" + df['Month'].astype(str).str.zfill(2)
        df['DisplayPeriod'] = df['MonthName'] + " " + df['Year'].astype(str)
    return df

with st.spinner("Synchronizing with Microsoft Fabric..."):
    if "current_brand" not in st.session_state:
        st.session_state.current_brand = None

    if st.session_state.current_brand != st.session_state.brand:
        raw_df = load_data(st.session_state.brand)
        st.session_state.original_df = raw_df.copy()
        st.session_state.current_df = raw_df.copy()
        st.session_state.current_brand = st.session_state.brand
        st.session_state.dirty = False

df = st.session_state.current_df.copy()
unique_periods = df[['PeriodSort', 'DisplayPeriod']].drop_duplicates().sort_values('PeriodSort', ascending=False)
period_list = unique_periods['DisplayPeriod'].tolist()

if df.empty:
    st.warning("No data retrieved from the database.")
    st.stop()

REVENUE_CLASSES = ['Net Sales', 'Other Income']

# =============================
# SIDEBAR NAVIGATION (ENHANCED UI ALIGNMENT)
# =============================
with st.sidebar:

    # -----------------------------
    # USER INFO
    # -----------------------------
    st.markdown(f"""
    <div class="sidebar-user-info">
        <p style='margin:0; font-size:11px; color:#64748B;'>SIGNED IN AS</p>
        <p style='margin:0; font-weight:700; color:#0F2044;'>👤 {st.session_state.logged_in_user.split('@')[0].title()}</p>
    </div>
    """, unsafe_allow_html=True)

    # -----------------------------
    # BRAND SELECTOR
    # -----------------------------
    st.markdown("<h2>Brand</h2>", unsafe_allow_html=True)

    selected_label = st.selectbox(
        "Select Brand",
        options=list(brand_map.keys()),
        index=list(brand_map.keys()).index(reverse_brand_map[st.session_state.brand]),
        label_visibility="visible"
    )

    selected_value = brand_map[selected_label]

    if selected_value != st.session_state.brand:
        st.session_state.brand = selected_value
        st.rerun()

    st.markdown("<hr>", unsafe_allow_html=True)

    # -----------------------------
    # NAVIGATION
    # -----------------------------
    view_options = ["📈 Financial Insights", "✏️ Ledger Editor"]
    # Hide Budget vs Actual for WED
    if st.session_state.brand == "pra":
        view_options.append("💰 Budget vs Actual")

    view_mode = st.radio(
        "Navigation",
        view_options,
        label_visibility="collapsed"
    )
    st.markdown("<hr>", unsafe_allow_html=True)

    # 📈 FINANCIAL INSIGHTS FILTERS
    if view_mode == "📈 Financial Insights":
        st.markdown("<h2>📊 Report Filters</h2>", unsafe_allow_html=True)
        date_range_type = st.radio(
            "Range Type",
            ["📅 Single/Multiple Months", "📆 Financial Year"],
            horizontal=True
        )

        if date_range_type == "📅 Single/Multiple Months":
            base_period = st.selectbox("Base Period", period_list)
            num_comparisons = st.number_input("Previous Periods", min_value=0, max_value=12, value=3)
            base_idx = period_list.index(base_period)
            selected_periods = period_list[base_idx: min(base_idx + num_comparisons + 1, len(period_list))]

        else:
            available_years = sorted(df['Year'].dropna().unique(), reverse=True)
            selected_year = st.selectbox("Financial Year (Apr–Mar)", available_years)
            fy_periods = get_financial_year_range(df, selected_year, start_month=4)
            if fy_periods:
                selected_periods = [p['display'] for p in fy_periods]
                st.info(f"{len(selected_periods)} periods: {fy_periods[0]['display']} → {fy_periods[-1]['display']}")
            else:
                selected_periods = []

        store_filter = st.selectbox(
            "🏢 Store",
            ["All"] + sorted(df['Store'].dropna().astype(str).unique())
        )

    # ✏️ LEDGER EDITOR
    elif view_mode == "✏️ Ledger Editor":

        st.markdown("<h2>📊 Report Filters</h2>", unsafe_allow_html=True)

        editor_date_range_type = st.radio(
            "Range Type",
            ["📅 Single/Multiple Months", "📆 Financial Year"],
            horizontal=True,
            key="editor_range_type"
        )

        if editor_date_range_type == "📅 Single/Multiple Months":
            editor_base_period = st.selectbox("Base Period", period_list, key="editor_base_period")
            editor_num_comparisons = st.number_input(
                "Number of Periods",
                min_value=1,
                max_value=12,
                value=3,
                key="editor_num_comparisons"
            )
            base_idx = period_list.index(editor_base_period)
            editor_selected_periods = period_list[base_idx: min(base_idx + editor_num_comparisons, len(period_list))]

        else:
            editor_available_years = sorted(df['Year'].dropna().unique(), reverse=True)
            editor_selected_year = st.selectbox(
                "Financial Year (Apr–Mar)",
                editor_available_years,
                key="editor_year"
            )
            fy_periods = get_financial_year_range(df, editor_selected_year, start_month=4)

            if fy_periods:
                editor_selected_periods = [p['display'] for p in fy_periods]
                st.info(f"{len(editor_selected_periods)} periods: {fy_periods[0]['display']} → {fy_periods[-1]['display']}")
            else:
                editor_selected_periods = []

        editor_store_filter = st.selectbox(
            "🏢 Store",
            ["All"] + sorted(df['Store'].dropna().astype(str).unique()),
            key="editor_store"
        )

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("<h2>✏️ Editor Filters</h2>", unsafe_allow_html=True)

        if editor_selected_periods:
            editor_filtered_df = df[df['DisplayPeriod'].isin(editor_selected_periods)].copy()

            if editor_store_filter != "All":
                editor_filtered_df = editor_filtered_df[editor_filtered_df['Store'] == editor_store_filter]

            if not editor_filtered_df.empty:
                editor_class = st.selectbox(
                    "Class",
                    ["All"] + sorted(editor_filtered_df['classification'].dropna().astype(str).unique()),
                    key="editor_class"
                )

                if editor_class != "All":
                    editor_filtered_df = editor_filtered_df[editor_filtered_df['classification'] == editor_class]

                if not editor_filtered_df.empty:
                    editor_account = st.selectbox(
                        "Account",
                        ["All"] + sorted(editor_filtered_df['account_name'].dropna().astype(str).unique()),
                        key="editor_account"
                    )

                    if editor_account != "All":
                        editor_filtered_df = editor_filtered_df[editor_filtered_df['account_name'] == editor_account]

                    if not editor_filtered_df.empty:
                        editor_partner = st.selectbox(
                            "Partner",
                            ["All"] + sorted(editor_filtered_df['partner_id_name'].dropna().astype(str).unique()),
                            key="editor_partner"
                        )

                        if editor_partner != "All":
                            editor_filtered_df = editor_filtered_df[editor_filtered_df['partner_id_name'] == editor_partner]

            st.session_state.editor_filtered_df = editor_filtered_df
            st.session_state.editor_selected_periods = editor_selected_periods
            st.session_state.editor_store_filter = editor_store_filter
        else:
            st.session_state.editor_filtered_df = pd.DataFrame()
            st.session_state.editor_selected_periods = []
            st.warning("No periods selected.")

    # 💰 BUDGET VS ACTUAL (ONLY PRA)
    elif view_mode == "💰 Budget vs Actual":
        st.markdown("<h2>📅 Budget Period</h2>", unsafe_allow_html=True)
        budget_base_period = st.selectbox(
            "Select Base Period",
            period_list,
            key="budget_base_period"
        )
        st.session_state.budget_selected_period = budget_base_period

    # Sign Out
    st.markdown("<hr>", unsafe_allow_html=True)
    if st.button("🚪 Sign Out", use_container_width=True, type="secondary"):
        st.session_state.logged_in = False
        st.session_state.logged_in_user = ""
        st.query_params.clear()
        st.cache_data.clear()
        st.rerun()

# =============================
# MAIN CONTENT
# =============================
st.markdown(f'<div class="main-header">💠 {view_mode.split(" ", 1)[1]}</div>', unsafe_allow_html=True)

# =============================
# FINANCIAL INSIGHTS VIEW
# =============================
if view_mode == "📈 Financial Insights":
    if not selected_periods:
        st.info("ℹ️ No periods selected or available.")
        st.stop()

    report_df = df[df['DisplayPeriod'].isin(selected_periods)].copy()
    if store_filter != "All": report_df = report_df[report_df['Store'] == store_filter]

    if report_df.empty:
        st.info("ℹ️ No data available for the selected filters.")
        st.stop()

    # --- KPI CARDS ---
    latest_period = selected_periods[0]
    latest_data = report_df[report_df['DisplayPeriod'] == latest_period]
    total_revenue = latest_data[latest_data['classification'].isin(REVENUE_CLASSES)]['Balance'].sum()
    total_expenses = latest_data[~latest_data['classification'].isin(REVENUE_CLASSES)]['Balance'].sum()
    net_profit = total_revenue - total_expenses
    profit_margin = (net_profit / total_revenue * 100) if total_revenue != 0 else 0

    cols = st.columns(4)
    kpi_configs = [
        (cols[0], "bg-revenue", "REVENUE", fmt_currency(total_revenue), f"{abs((total_expenses/total_revenue*100) if total_revenue!=0 else 0):.1f}% OPEX"),
        (cols[1], "bg-expense", "EXPENSES", fmt_currency(total_expenses), f"{abs((total_expenses/total_revenue*100) if total_revenue!=0 else 0):.1f}% of Rev"),
        (cols[2], "bg-profit", "NET PROFIT", fmt_currency(net_profit), f"{abs((total_expenses/total_revenue*100) if total_revenue!=0 else 0):.1f}% Margin"),
        (cols[3], "bg-margin", "MARGIN", f"{profit_margin:.1f}%", f"{latest_period[:3]}")
    ]
    for col, bg_class, title, value, sub in kpi_configs:
        col.markdown(f"""
        <div class="kpi-container {bg_class}">
            <div class="kpi-title">{title}</div>
            <div class="kpi-value">{value}</div>
            <div><span class="kpi-sub">{sub}</span></div>
        </div>
        """, unsafe_allow_html=True)

    st.write("<br>", unsafe_allow_html=True)

    # --- CHART ---
    profit_df = calculate_profit_metrics(report_df, selected_periods, REVENUE_CLASSES)
    fig = go.Figure()
    fig.add_trace(go.Bar(x=profit_df['DisplayPeriod'], y=profit_df['Revenue'], name='Revenue', marker_color='#0AB370', opacity=0.8))
    fig.add_trace(go.Bar(x=profit_df['DisplayPeriod'], y=profit_df['Expenses'], name='Expenses', marker_color='#F43F5E', opacity=0.8))
    fig.add_trace(go.Scatter(x=profit_df['DisplayPeriod'], y=profit_df['Profit'], mode='lines+markers', name='Net Profit', line=dict(color='#2563EB', width=3, shape='spline'), marker=dict(size=8, color='#1D4ED8'), yaxis='y'))
    fig.update_layout(title=dict(text='<b>Revenue, Expenses & Net Profit</b>', font=dict(size=18, color='#0F2044')), barmode='group', plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', margin=dict(l=40, r=40, t=50, b=40), yaxis=dict(gridcolor='#E2E8F0', title='Amount (₹)', tickformat='₹,.0f'), xaxis=dict(showline=True, linecolor='#CBD5E1', tickfont=dict(size=10)), hovermode='x unified', legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1, bgcolor='rgba(255,255,255,0.9)', font=dict(size=10)))
    st.plotly_chart(fig, width='stretch')

    st.markdown("---")
    
    hierarchy = build_hierarchy_data(report_df, selected_periods)

    col1, col2 = st.columns([3, 2])
    with col1:
        st.markdown("### 📋 P&L Statement")
    with col2:
        export_col1, export_col2, export_col3 = st.columns([1, 1, 1])
        with export_col1:
            if st.button("⊞ Expand All", key="pnl_expand_btn", width='stretch'):
                st.session_state.expand_pnl = True  
                st.rerun()
        with export_col2:
            if st.button("Collapse All", key="pnl_collapse_btn", width='stretch'):
                st.session_state.expand_pnl = False 
                st.rerun()
        with export_col3:
            brand_name = reverse_brand_map.get(st.session_state.brand, "Unknown")

            excel_file = build_excel_report(
                hierarchy=hierarchy, 
                periods=selected_periods, 
                store_filter=store_filter,
                brand=st.session_state.brand,
                expand_all=st.session_state.get('expand_pnl', False),
                open_classifications=st.session_state.open_classifications,
                open_accounts=st.session_state.open_accounts
            )
            
            st.download_button(
                label="📥 Export",
                data=excel_file,
                file_name=f"{brand_name}_PnL_Statement_{store_filter}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="pnl_export_main_btn", 
                use_container_width=True
            )

    # ==========================================
    # HIGH-PERFORMANCE HTML/CSS GRID TABLE
    # ==========================================
    num_periods = len(selected_periods)
    grid_template = f"350px repeat({num_periods}, minmax(130px, 1fr)) minmax(140px, 1fr)"

    pnl_open_attr = "open" if st.session_state.expand_pnl else ""
    
    html_parts = []
    html_parts.append(f"""
    <style>
    .pnl-container {{ width: 100%; overflow-x: auto; border: 1px solid #CBD5E1; border-radius: 8px; background: #FFFFFF; padding-bottom: 10px; }}
    .pnl-table-wrapper {{ min-width: 100%; width: max-content; display: flex; flex-direction: column; }}
    .pnl-row {{ 
        display: grid; 
        grid-template-columns: {grid_template}; 
        border-bottom: 1px solid #E2E8F0; 
        align-items: center; 
        transition: background 0.2s; 
        width: 100%;
        background-color: inherit;
    }}
    .pnl-row:hover {{ background-color: #F8FAFC; }}
    .pnl-header {{ background-color: #0F2044 !important; color: white; font-weight: bold; position: sticky; top: 0; z-index: 10; font-size: 12px; letter-spacing: 0.5px; }}
    .pnl-cell {{ padding: 10px 12px; white-space: nowrap; font-size: 13px; }}
    .pnl-cell:first-child {{ 
        position: sticky; 
        left: 0; 
        background-color: inherit; 
        z-index: 5; 
        border-right: 1px solid #E2E8F0; 
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
    }}
    .pnl-header .pnl-cell:first-child {{ background-color: #0F2044; z-index: 15; }}
    .align-right {{ text-align: right; }}
    .val-pos {{ color: #059669; font-weight: 600; font-family: 'DM Mono', monospace; }}
    .val-neg {{ color: #E11D48; font-weight: 600; font-family: 'DM Mono', monospace; }}
    .val-tot {{ color: #0F2044; font-weight: 800; font-family: 'DM Mono', monospace; background: #EEF2F9; border-radius: 4px; padding: 3px 6px; display: inline-block; }}
    .lvl-1 {{ font-weight: 700; color: #0F2044; background: #F1F5F9; cursor: pointer; }}
    .lvl-2 {{ font-weight: 600; color: #334155; background: #FFFFFF; cursor: pointer; }}
    .lvl-3 {{ color: #64748B; background: #FFFFFF; }}
    details > summary {{ list-style: none; outline: none; }}
    details > summary::-webkit-details-marker {{ display: none; }}
    .arrow {{ display: inline-block; width: 18px; font-size: 11px; transition: transform 0.2s; color: #64748B; margin-right: 4px; }}
    details[open] > summary .arrow {{ transform: rotate(90deg); color: #0F2044; }}
    </style>
    
    <div class='pnl-container'>
        <div class='pnl-table-wrapper'>
    """)

    html_parts.append("<div class='pnl-row pnl-header'>")
    html_parts.append("<div class='pnl-cell'>PARTICULARS</div>")
    for p in selected_periods:
        short_p = p[:3] + " " + p[-2:]
        html_parts.append(f"<div class='pnl-cell align-right'>{short_p.upper()}</div>")
    html_parts.append("<div class='pnl-cell align-right'>TOTAL</div></div>")

    # FIX: Use pnl specific state
    pnl_open_attr = "open" if st.session_state.get('expand_pnl', False) else ""

    for cls_name, cls_data in hierarchy.items():
        html_parts.append(f"<details {pnl_open_attr}><summary class='pnl-row lvl-1'>")
        html_parts.append(f"<div class='pnl-cell' title='{cls_name.upper()}'><span class='arrow'>▶</span> 📂 {cls_name.upper()}</div>")
        for p in selected_periods:
            v = cls_data['totals'].get(p, 0)
            color_cls = "val-pos" if v >= 0 else "val-neg"
            html_parts.append(f"<div class='pnl-cell align-right {color_cls}'>{fmt_currency(v)}</div>")
        
        cls_tot = sum(cls_data['totals'].values())
        html_parts.append(f"<div class='pnl-cell align-right'><span class='val-tot'>{fmt_currency(cls_tot)}</span></div>")
        html_parts.append("</summary>")

        for acc_name, acc_data in cls_data['accounts'].items():
            html_parts.append(f"<details {pnl_open_attr}><summary class='pnl-row lvl-2'>")
            html_parts.append(f"<div class='pnl-cell' title='{acc_name}' style='padding-left: 28px;'><span class='arrow'>▶</span> 📄 {acc_name}</div>")
            for p in selected_periods:
                v = acc_data['totals'].get(p, 0)
                color_cls = "val-pos" if v >= 0 else "val-neg"
                html_parts.append(f"<div class='pnl-cell align-right {color_cls}'>{fmt_currency(v)}</div>")
            
            acc_tot = sum(acc_data['totals'].values())
            color_cls = "val-pos" if acc_tot >= 0 else "val-neg"
            html_parts.append(f"<div class='pnl-cell align-right {color_cls}' style='font-weight: 800;'>{fmt_currency(acc_tot)}</div>")
            html_parts.append("</summary>")

            for partner_name, prt_totals in acc_data['partners'].items():
                html_parts.append("<div class='pnl-row lvl-3'>")
                html_parts.append(f"<div class='pnl-cell' title='{partner_name}' style='padding-left: 65px;'>• {partner_name}</div>")
                for p in selected_periods:
                    v = prt_totals.get(p, 0)
                    color_cls = "val-pos" if v >= 0 else "val-neg"
                    html_parts.append(f"<div class='pnl-cell align-right {color_cls}'>{fmt_currency(v)}</div>")
                
                prt_tot = sum(prt_totals.values())
                color_cls = "val-pos" if prt_tot >= 0 else "val-neg"
                html_parts.append(f"<div class='pnl-cell align-right {color_cls}'>{fmt_currency(prt_tot)}</div>")
                html_parts.append("</div>")

            html_parts.append("</details>") 
        html_parts.append("</details>") 

    html_parts.append("</div></div>") 
    st.markdown("".join(html_parts), unsafe_allow_html=True)
    st.write("<br>", unsafe_allow_html=True)
    st.markdown("---")

    # 1. Prepare Data for Store Comparison
    # Check if a full year is selected (Assuming 12 months = Full FY)
    is_full_fy = len(selected_periods) >= 12 

    if is_full_fy:
        # IF FULL FY: Use all data (this aggregates all 12 months per store)
        comp_df = report_df.copy()
        display_period_label = "Financial Year"
    else:
        # ELSE: Use only the first selected month (Base Period)
        base_period = selected_periods[0]
        comp_df = report_df[report_df['DisplayPeriod'] == base_period].copy()
        display_period_label = base_period

    relevant_stores = sorted(comp_df['Store'].unique().tolist())
    store_hierarchy = build_hierarchy_data(comp_df, relevant_stores, group_by='Store')

    # 2. UI Header & Independent Buttons
    col1, col2 = st.columns([3, 2])
    with col1:
        st.markdown(f"### 🏪 Store Comparison ({display_period_label})")
    with col2:
        s_exp1, s_exp2, s_exp3 = st.columns([1, 1, 1])
        with s_exp1:
            if st.button("⊞ Expand All", key="btn_store_expand", width='stretch'):
                st.session_state.expand_store = True
                st.rerun()
        with s_exp2:
            if st.button("Collapse All", key="btn_store_collapse", width='stretch'):
                st.session_state.expand_store = False
                st.rerun()
        with s_exp3:
            brand_name = reverse_brand_map.get(st.session_state.brand, "Unknown")

            store_excel = build_excel_report(
                hierarchy=store_hierarchy, 
                periods=relevant_stores, 
                store_filter="Comparison", 
                report_type="store",
                expand_all=st.session_state.get('expand_store', False)
            )
            st.download_button(
                label="📥 Export",
                type="primary",
                data=store_excel,
                file_name=f"{brand_name}_Store_Comparison_{display_period_label}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_store_export",
                use_container_width=True
            )

    # 3. Render Table with Independent State
    # FIX: Use store specific state
    store_open_attr = "open" if st.session_state.expand_store else ""

    num_stores = len(relevant_stores)
    store_grid_template = f"350px repeat({num_stores}, 150px) 150px"

    store_html = []
    store_html.append(f"""
    <style>
    /* Container handles the scroll */
    .store-container {{ width: 100%; overflow-x: auto; border: 1px solid #CBD5E1; border-radius: 8px; background: #FFFFFF; padding-bottom: 10px; }}
    
    /* Wrapper guarantees background stretches 100% */
    .store-table-wrapper {{ min-width: 100%; width: max-content; display: flex; flex-direction: column; }}
    
    /* Rows span exactly the wrapper's width */
    .store-row {{ 
        display: grid; 
        grid-template-columns: {store_grid_template}; 
        border-bottom: 1px solid #E2E8F0; 
        align-items: center; 
        transition: background 0.2s; 
        width: 100%;
        background-color: inherit;
    }}
    .store-row:hover {{ background-color: #F8FAFC; }}
    
    /* Header styling */
    .store-header {{ background-color: #0F2044 !important; color: white; font-weight: bold; position: sticky; top: 0; z-index: 10; font-size: 12px; letter-spacing: 0.5px; }}
    .store-cell {{ padding: 10px 12px; white-space: nowrap; font-size: 13px; }}
    
    /* Sticky first column fix: background-color must be solid, not inherit, to prevent overlap */
    .store-cell:first-child {{ 
        position: sticky; 
        left: 0; 
        background-color: #FFFFFF; 
        z-index: 5; 
        border-right: 1px solid #E2E8F0; 
        overflow: hidden;
        text-overflow: ellipsis;
        white-space: nowrap;
    }}
    .store-header .store-cell:first-child {{ background-color: #0F2044; z-index: 15; }}
    
    .align-right {{ text-align: right; }}
    
    /* Value Typography */
    .val-pos {{ color: #059669; font-weight: 600; font-family: 'DM Mono', monospace; }}
    .val-neg {{ color: #E11D48; font-weight: 600; font-family: 'DM Mono', monospace; }}
    .val-tot {{ color: #0F2044; font-weight: 800; font-family: 'DM Mono', monospace; background: #EEF2F9; border-radius: 4px; padding: 3px 6px; display: inline-block; }}
    
    /* Hierarchical Styling */
    .lvl-1 {{ font-weight: 700; color: #0F2044; background: #F1F5F9; cursor: pointer; }}
    .lvl-2 {{ font-weight: 600; color: #334155; background: #FFFFFF; cursor: pointer; }}
    .lvl-3 {{ color: #64748B; background: #FFFFFF; }}
    
    /* Accordion Details/Summary Reset */
    details > summary {{ list-style: none; outline: none; }}
    details > summary::-webkit-details-marker {{ display: none; }}
    .arrow {{ display: inline-block; width: 18px; font-size: 11px; transition: transform 0.2s; color: #64748B; margin-right: 4px; }}
    details[open] > summary .arrow {{ transform: rotate(90deg); color: #0F2044; }}
    </style>
    
    <div class='store-container'>
        <div class='store-table-wrapper'>
    """)

    store_html.append("<div class='store-row store-header'><div class='store-cell'>PARTICULARS</div>")
    for s in relevant_stores:
        store_html.append(f"<div class='store-cell align-right'>{s.upper()}</div>")
    store_html.append("<div class='store-cell align-right'>TOTAL</div></div>")

    for cls_name, cls_data in store_hierarchy.items():
        store_html.append(f"<details {store_open_attr}><summary class='store-row lvl-1'>")
        store_html.append(f"<div class='store-cell'><span class='arrow'>▶</span> 📂 {cls_name.upper()}</div>")
        for s in relevant_stores:
            v = cls_data['totals'].get(s, 0)
            store_html.append(f"<div class='store-cell align-right {'val-pos' if v >= 0 else 'val-neg'}'>{fmt_currency(v)}</div>")
        store_html.append(f"<div class='store-cell align-right'><span class='val-tot'>{fmt_currency(sum(cls_data['totals'].values()))}</span></div>")
        store_html.append("</summary>")

        for acc_name, acc_data in cls_data['accounts'].items():
            store_html.append(f"<details {store_open_attr}><summary class='store-row lvl-2'>")
            store_html.append(f"<div class='store-cell' style='padding-left: 28px;'><span class='arrow'>▶</span> 📄 {acc_name}</div>")
            for s in relevant_stores:
                v = acc_data['totals'].get(s, 0)
                store_html.append(f"<div class='store-cell align-right {'val-pos' if v >= 0 else 'val-neg'}'>{fmt_currency(v)}</div>")
            store_html.append(f"<div class='store-cell align-right' style='font-weight: 800;'>{fmt_currency(sum(acc_data['totals'].values()))}</div>")
            store_html.append("</summary>")

            for partner_name, prt_totals in acc_data['partners'].items():
                store_html.append("<div class='store-row lvl-3'>")
                store_html.append(f"<div class='store-cell' style='padding-left: 65px;'>• {partner_name}</div>")
                for s in relevant_stores:
                    v = prt_totals.get(s, 0)
                    store_html.append(f"<div class='store-cell align-right {'val-pos' if v >= 0 else 'val-neg'}'>{fmt_currency(v)}</div>")
                store_html.append(f"<div class='store-cell align-right'>{fmt_currency(sum(prt_totals.values()))}</div>")
                store_html.append("</div>")

            store_html.append("</details>") 
        store_html.append("</details>") 

    store_html.append("</div></div>")
    st.markdown("".join(store_html), unsafe_allow_html=True)

# ===========================
# LEDGER EDITOR VIEW
# ===========================
elif view_mode == "✏️ Ledger Editor":
    if not hasattr(st.session_state, 'editor_filtered_df') or st.session_state.editor_filtered_df.empty:
        st.info("ℹ️ Please select periods and filters in the sidebar to view/edit data.")
        st.stop()
    
    if "reset_editor" not in st.session_state:
        st.session_state.reset_editor = 0

    editor_df = st.session_state.editor_filtered_df
    editor_periods = st.session_state.editor_selected_periods
    editor_store = st.session_state.editor_store_filter
    
    filter_summary = f"**Current View:** {len(editor_periods)} period(s): {', '.join(editor_periods)}"
    if editor_store != "All":
        filter_summary += f" | Store: {editor_store}"
    st.markdown(filter_summary)
    
    st.markdown("Double-click any period balance to edit. Changes will be reflected in the aggregated total.")

    dimension_columns = ['Store', 'classification', 'account_name', 'partner_id_name']
    
    if editor_df.empty:
        st.warning("No records found for the selected filters.")
    else:
        pivot_df = editor_df.groupby(dimension_columns + ['DisplayPeriod'], as_index=False)['Balance'].sum()
        
        pivot_df = pivot_df.pivot_table(
            index=dimension_columns,
            columns='DisplayPeriod',
            values='Balance',
            aggfunc='sum',
            fill_value=0.0
        ).reset_index()
        
        pivot_df.columns.name = None
        period_columns = [col for col in pivot_df.columns if col not in dimension_columns]
        
        def get_period_sort_key(period):
            try:
                month_name, year = period.rsplit(' ', 1)
                month_map = {
                    'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
                    'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12
                }
                month_num = month_map.get(month_name, 0)
                return (int(year), month_num)
            except:
                return (9999, 99)
        
        if period_columns:
            period_columns = sorted(period_columns, key=get_period_sort_key)
            pivot_df['Total'] = pivot_df[period_columns].sum(axis=1)
            column_order = dimension_columns + period_columns + ['Total']
            pivot_df = pivot_df[column_order]
            
            for col in period_columns + ['Total']:
                pivot_df[col] = pd.to_numeric(pivot_df[col], errors='coerce').fillna(0.0).apply(fmt_currency)
            
            pivot_df = pivot_df.sort_values(['classification', 'account_name', 'partner_id_name'])
            
            st.markdown(f"""
            <span style='color:#64748B; font-size:0.8rem; font-weight:500;'>
            {len(pivot_df)} unique account-partner combinations | 
            Showing balances for {len(period_columns)} periods: {period_columns[0]} → {period_columns[-1]}
            </span>
            """, unsafe_allow_html=True)

            column_config = {
                "Store": st.column_config.TextColumn("Store", disabled=True, width="stretch"),
                "classification": st.column_config.TextColumn("Classification", disabled=True),
                "account_name": st.column_config.TextColumn("Account", disabled=True),
                "partner_id_name": st.column_config.TextColumn("Partner", disabled=True),
                "Total": st.column_config.TextColumn("Total (All Periods)", disabled=True)
            }
            
            for period in period_columns:
                try:
                    month_name, year = period.rsplit(' ', 1)
                    short_period = month_name[:3] + " " + year[-2:]
                except:
                    short_period = period
                
                column_config[period] = st.column_config.TextColumn(
                    short_period, required=True
                )

            filtered_pivot_df = pivot_df.copy()
            if 'editor_class' in st.session_state and st.session_state.editor_class != "All":
                filtered_pivot_df = filtered_pivot_df[filtered_pivot_df['classification'] == st.session_state.editor_class]
            if 'editor_account' in st.session_state and st.session_state.editor_account != "All":
                filtered_pivot_df = filtered_pivot_df[filtered_pivot_df['account_name'] == st.session_state.editor_account]
            if 'editor_partner' in st.session_state and st.session_state.editor_partner != "All":
                filtered_pivot_df = filtered_pivot_df[filtered_pivot_df['partner_id_name'] == st.session_state.editor_partner]

            current_editor_key = f"editor_{st.session_state.reset_editor}"
            editor_df_widget = st.data_editor(
                filtered_pivot_df, 
                key=current_editor_key, 
                width='stretch',
                hide_index=True,
                num_rows="fixed",
                disabled=['Store', 'classification', 'account_name', 'partner_id_name', 'Total'],
                column_config=column_config
            )

            changes_summary = []
            if current_editor_key in st.session_state:
                edits = st.session_state[current_editor_key].get("edited_rows", {})
                for row_idx, changed_cols in edits.items():
                    row_data = filtered_pivot_df.iloc[row_idx]
                    for period, new_val in changed_cols.items():
                        if period in period_columns:
                            clean_val = str(new_val).replace(',', '').replace('₹','')
                            changes_summary.append({
                                'store': row_data['Store'],
                                'classification': row_data['classification'],
                                'account': row_data['account_name'],
                                'partner': row_data['partner_id_name'],
                                'period': period,
                                'new_value': float(clean_val)
                            })

            st.session_state.dirty = len(changes_summary) > 0

            st.write("<br>", unsafe_allow_html=True)
            col1, col2, col3 = st.columns([6, 2, 2])

            with col1:
                if st.session_state.dirty:
                    st.info(f"📝 {len(changes_summary)} pending changes.")

            with col2:
                if st.button("🗑️ Discard", width='stretch', disabled=not st.session_state.dirty):
                    if current_editor_key in st.session_state:
                        del st.session_state[current_editor_key]
                    st.session_state.reset_editor += 1
                    st.session_state.dirty = False
                    st.rerun()

            with col3:
                if st.button("💾 Save to Fabric", width='stretch', disabled=not st.session_state.dirty, type="primary"):
                    with st.spinner("Syncing changes with Fabric..."):
                        success_count, error_count = 0, 0
                        month_map = {"Jan": "January", "Feb": "February", "Mar": "March", "Apr": "April", "May": "May", "Jun": "June", "Jul": "July", "Aug": "August", "Sep": "September", "Oct": "October", "Nov": "November", "Dec": "December"}
                        current_time_fabric = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

                        for change in changes_summary:
                            p_parts = change['period'].split()
                            full_month = month_map.get(p_parts[0], p_parts[0])
                            raw_year = p_parts[1]
                            full_year = 2000 + int(raw_year) if len(raw_year) == 2 else int(raw_year)
                            
                            variables = {
                                "year": full_year,
                                "monthName": full_month,
                                "store": str(change['store']).strip(),
                                "balance": float(change['new_value']),
                                "account_name": str(change['account']).strip(),
                                "classification": str(change['classification']).strip(),
                                "partner_id_name": str(change['partner']).strip(),
                                "last_modified_at": current_time_fabric,
                                "last_modified_user": st.session_state.get('logged_in_user', 'Unknown')
                            }

                            try:
                                response = run_graphql(UPSERT_MUTATION, variables)
                                if response and "errors" not in response:
                                    success_count += 1
                                else:
                                    error_msg = response['errors'][0]['message'] if response else "No response"
                                    st.error(f"Failed to save {change['account']}: {error_msg}")
                                    error_count += 1
                            except Exception as e:
                                st.error(f"Connection error: {str(e)}")
                                error_count += 1

                        if error_count == 0:
                            st.success(f"✅ Successfully synced {success_count} records to Fabric!")
                            load_data.clear() 
                            if current_editor_key in st.session_state:
                                del st.session_state[current_editor_key]
                            st.session_state.reset_editor += 1
                            st.session_state.dirty = False
                            st.rerun()
                        else:
                            st.warning(f"Process complete: {success_count} saved, {error_count} failed.")

# ===========================
# BUDGETING VIEW
# ===========================
elif view_mode == "💰 Budget vs Actual":
    if "budget_base_period" not in st.session_state:
        st.info("ℹ️ Please select a Base Period in the sidebar to view budgeting data.")
        st.stop()

    base_period = st.session_state.budget_base_period

    classification_order = [
        'Net Sales', 'Other Income', 'Cost of Goods Sold (COGS)', 'Employee cost',
        'Rent and Utilities', 'Marketing and Advertisment', 'Admin Expenses',
        'Logistics', 'Other Expenses', 'Finance cost', 'Supplier Payments',
        'Purchase Expense', 'Depreciation'
    ]

    st.markdown("""
    <style>
        [data-testid="stDataFrame"] th p {
            font-weight: 800 !important;
            text-transform: uppercase !important;
            text-align: center !important;
            color: #1E293B !important;
        }
        .super-kpi-card {
            border-radius: 12px;
            padding: 22px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.05);
            margin-bottom: 20px;
        }
        
        .card-revenue { 
            background: linear-gradient(90deg, #F0FDF4 0%, #FFFFFF 100%);
            border: 1px solid #DCFCE7;
            border-left: 6px solid #22C55E !important; 
        } 
        
        .card-expense { 
            background: linear-gradient(90deg, #FEF2F2 0%, #FFFFFF 100%);
            border: 1px solid #FEE2E2;
            border-left: 6px solid #EF4444 !important; 
        } 
        
        .super-kpi-header {
            font-size: 0.8rem;
            font-weight: 700;
            color: #475569;
            text-transform: uppercase;
            letter-spacing: 0.8px;
            margin-bottom: 18px;
        }
        .partition-container {
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .partition {
            flex: 1;
            text-align: center;
            border-right: 1px solid rgba(0,0,0,0.05);
        }
        .partition:last-child { border-right: none; }
        
        .partition-label { font-size: 0.65rem; color: #64748B; font-weight: 700; margin-bottom: 6px; }
        .partition-value { font-size: 1.15rem; font-weight: 700; color: #0F172A; font-family: 'DM Mono', monospace; }
        
        .text-revenue { color: #15803D !important; font-weight: 800; }
        .text-expense { color: #B91C1C !important; font-weight: 800; }
        .text-neutral { color: #854D0E !important; background: #FEF9C3; padding: 2px 8px; border-radius: 4px; }
    </style>
    """, unsafe_allow_html=True)
    
    BUDGET_QUERY = "query budgetData { executesp_pr_readBudgetData { Particulars Month Year Budget } }"
    
    @st.cache_data(ttl=600)
    def fetch_budget_data():
        response = run_graphql(BUDGET_QUERY)
        return pd.DataFrame(response["data"]["executesp_pr_readBudgetData"]) if response else pd.DataFrame()

    raw_budget_df = fetch_budget_data()
    actual_data = df[df['DisplayPeriod'] == base_period].copy()
    if 'editor_store_filter' in st.session_state and st.session_state.editor_store_filter != "All":
        actual_data = actual_data[actual_data['Store'] == st.session_state.editor_store_filter]
    
    actual_summary = actual_data.groupby('classification')['Balance'].sum().reset_index()
    actual_summary.rename(columns={'classification': 'Particulars', 'Balance': 'Actual'}, inplace=True)

    if not raw_budget_df.empty:
        raw_budget_df['DisplayPeriod'] = raw_budget_df['Month'] + " " + raw_budget_df['Year'].astype(str)
        period_budget_df = raw_budget_df[raw_budget_df['DisplayPeriod'] == base_period]
        
        # Safely handle periods with zero budget data
        if not period_budget_df.empty:
            budget_pivot = period_budget_df.pivot_table(index='Particulars', values='Budget', aggfunc='sum').reset_index()
            pivot_df = pd.merge(pd.DataFrame({'Particulars': classification_order}), budget_pivot, on='Particulars', how='left').fillna(0.0)
        else:
            pivot_df = pd.DataFrame({'Particulars': classification_order, 'Budget': 0.0})
    else:
        pivot_df = pd.DataFrame({'Particulars': classification_order, 'Budget': 0.0})

    # Merge actuals
    pivot_df = pd.merge(pivot_df, actual_summary, on='Particulars', how='left').fillna(0.0)

    # Failsafe: Ensure 'Actual' column exists even if actual_summary was completely empty
    if 'Actual' not in pivot_df.columns:
        pivot_df['Actual'] = 0.0

    # ==========================================

    def get_val(df, part, col):
        return df.loc[df['Particulars'] == part, col].sum()

    rev_b = get_val(pivot_df, 'Net Sales', 'Budget') + get_val(pivot_df, 'Other Income', 'Budget')
    rev_a = get_val(pivot_df, 'Net Sales', 'Actual') + get_val(pivot_df, 'Other Income', 'Actual')
    exp_b = get_val(pivot_df, 'Cost of Goods Sold (COGS)', 'Budget')
    exp_a = get_val(pivot_df, 'Cost of Goods Sold (COGS)', 'Actual')
    op_exp_b = (get_val(pivot_df, 'Employee cost', 'Budget') + 
                get_val(pivot_df, 'Rent and Utilities', 'Budget') + 
                get_val(pivot_df, 'Marketing and Advertisment', 'Budget') + 
                get_val(pivot_df, 'Admin Expenses', 'Budget') + 
                get_val(pivot_df, 'Logistics', 'Budget') +
                get_val(pivot_df, 'Other Expenses', 'Budget')
    )
    op_exp_a = (get_val(pivot_df, 'Employee cost', 'Actual') + 
                get_val(pivot_df, 'Rent and Utilities', 'Actual') + 
                get_val(pivot_df, 'Marketing and Advertisment', 'Actual') + 
                get_val(pivot_df, 'Admin Expenses', 'Actual') + 
                get_val(pivot_df, 'Logistics', 'Actual') +
                get_val(pivot_df, 'Other Expenses', 'Actual')
    )
    fin_cost_b = get_val(pivot_df, 'Finance cost', 'Budget')
    fin_cost_a = get_val(pivot_df, 'Finance cost', 'Actual')

    final_rows = [
        'Net Sales', 'Other Income',
        {'Particulars': 'TOTAL REVENUE', 'Budget': rev_b, 'Actual': rev_a, 'is_calc': True},
        'Cost of Goods Sold (COGS)',
        {'Particulars': 'TOTAL EXPENSE', 'Budget': exp_b, 'Actual': exp_a, 'is_calc': True},
        {'Particulars': 'GROSS PROFIT', 'Budget': rev_b - exp_b, 'Actual': rev_a - exp_a, 'is_calc': True},
        'Employee cost', 'Rent and Utilities', 'Marketing and Advertisment', 
        'Admin Expenses', 'Logistics', 'Other Expenses',
        {'Particulars': 'TOTAL OPERATING EXPENSE', 'Budget': op_exp_b, 'Actual': op_exp_a, 'is_calc': True},
        {'Particulars': 'OPERATING PROFIT (EBIT)', 'Budget': rev_b - exp_b - op_exp_b, 'Actual': rev_a - exp_a - op_exp_a, 'is_calc': True},
        'Finance cost', 'Depreciation',
        {'Particulars': 'PBT', 'Budget': rev_b - exp_b - op_exp_b - fin_cost_b, 'Actual': rev_a - exp_a - op_exp_a - fin_cost_a, 'is_calc': True},
        'Supplier Payments', 'Purchase Expense'
    ]

    processed_data = []
    for item in final_rows:
        if isinstance(item, dict):
            processed_data.append(item)
        else:
            row_match = pivot_df[pivot_df['Particulars'] == item]
            b_val = row_match['Budget'].values[0] if not row_match.empty else 0.0
            a_val = row_match['Actual'].values[0] if not row_match.empty else 0.0
            processed_data.append({'Particulars': item, 'Budget': b_val, 'Actual': a_val, 'is_calc': False})

    table_df = pd.DataFrame(processed_data)
    table_df['Change %'] = table_df.apply(lambda r: ((r['Actual'] - r['Budget']) / r['Budget'] * 100) if r['Budget'] != 0 else (100.0 if r['Actual'] > 0 else 0.0), axis=1)

    st.subheader(f"🎯 Performance Analytics: {base_period}")

    rev_diff = rev_a - rev_b
    rev_var_pct = (rev_diff / rev_b * 100) if rev_b != 0 else 0
    rev_text_class = "text-revenue" if rev_diff > 0 else "text-expense" if rev_diff < 0 else "text-neutral"

    total_planned_exp = op_exp_b + exp_b + fin_cost_b
    total_actual_exp = op_exp_a + exp_a + fin_cost_a
    exp_diff = total_actual_exp - total_planned_exp
    exp_var_pct = (exp_diff / total_planned_exp * 100) if total_planned_exp != 0 else 0
    exp_text_class = "text-expense" if exp_diff > 0 else "text-revenue" if exp_diff < 0 else "text-neutral"

    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown(f"""
        <div class="super-kpi-card card-revenue">
            <div class="super-kpi-header">📊 Revenue Performance</div>
            <div class="partition-container">
                <div class="partition">
                    <div class="partition-label">TARGET REVENUE</div>
                    <div class="partition-value">{fmt_currency(rev_b)}</div>
                </div>
                <div class="partition">
                    <div class="partition-label">ACTUAL REVENUE</div>
                    <div class="partition-value">{fmt_currency(rev_a)}</div>
                </div>
                <div class="partition">
                    <div class="partition-label">VARIANCE %</div>
                    <div class="partition-value {rev_text_class}">{rev_var_pct:+.1f}%</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with col_right:
        st.markdown(f"""
        <div class="super-kpi-card card-expense">
            <div class="super-kpi-header">💸 Budget Utilization</div>
            <div class="partition-container">
                <div class="partition">
                    <div class="partition-label">PLANNED BUDGET</div>
                    <div class="partition-value">{fmt_currency(total_planned_exp)}</div>
                </div>
                <div class="partition">
                    <div class="partition-label">EXPENDITURE TO DATE</div>
                    <div class="partition-value">{fmt_currency(total_actual_exp)}</div>
                </div>
                <div class="partition">
                    <div class="partition-label">VARIANCE %</div>
                    <div class="partition-value {exp_text_class}">{exp_var_pct:+.1f}%</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.write("<br>", unsafe_allow_html=True)

    def style_rows(row):
        styles = [''] * len(row)
        v = row['Change %']
        if v > 0: styles[row.index.get_loc('Change %')] = 'background-color: #FEF2F2; color: #991B1B;'
        elif v < 0: styles[row.index.get_loc('Change %')] = 'background-color: #F0FDF4; color: #166534;'
        else: styles[row.index.get_loc('Change %')] = 'background-color: #FEF9C3; color: #854D0E;'
        
        if row['is_calc']:
            return ['font-weight: 800; background-color: #F8FAFC; color: #1E293B; border-top: 2px solid #334155;'] * len(row)
        return styles

    styler = table_df.style.apply(style_rows, axis=1)
    styler = styler.format({
        'Budget': lambda x: fmt_currency(x),
        'Actual': lambda x: fmt_currency(x),
        'Change %': '{:+.1f}%'
    })

    st.dataframe(
        styler,
        width='stretch',
        hide_index=True,
        height=(len(table_df) * 36) + 40,
        column_config={
            "Particulars": st.column_config.TextColumn("PARTICULARS", width="large"),
            "Budget": st.column_config.TextColumn("BUDGET", width="medium"),
            "Actual": st.column_config.TextColumn("ACTUAL", width="medium"),
            "Change %": st.column_config.TextColumn("CHANGE %", width="small"),
            "is_calc": None
        }
    )
